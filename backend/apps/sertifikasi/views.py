from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import get_object_or_404

from .models import Sertifikasi
from .serializers import SertifikasiSerializer
from ..authentication.permissions import isAdminPermission

from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListAPIView
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes

from url_filter.integrations.drf import DjangoFilterBackend

from datetime import datetime
from openpyxl import Workbook

class SertifikasiViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)

    def get_all_objects(self, request):
        if request.user.is_admin:
            queryset = Sertifikasi.objects.all()
        else:
            queryset = Sertifikasi.objects.filter(is_validated=True)
        return queryset

    def get_permissions(self):
        if self.action == 'destroy':
            permission_classes = [isAdminPermission]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def list(self, request):
        queryset = self.get_all_objects(request)
        serializer = SertifikasiSerializer(queryset, many=True)
        return Response(serializer.data)
    
    def create(self, request):
        serializer = SertifikasiSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def retrieve(self, request, pk=None):
        queryset = self.get_all_objects(request)
        sertifikasi = get_object_or_404(queryset, pk=pk)
        serializer = SertifikasiSerializer(sertifikasi)
        return Response(serializer.data)

    def update(self, request, pk=None):
        pass

    def partial_update(self, request, pk=None):
        pass

    def destroy(self, request, pk=None):
        queryset = Sertifikasi.objects.all()
        sertifikasi = get_object_or_404(queryset, pk=pk)
        sertifikasi.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

# Validated API
@api_view(['POST'])
@permission_classes([isAdminPermission])
def set_validate(request, pk):
    if request.method == 'POST':
        queryset = Sertifikasi.objects.all()
        sertifikasi = get_object_or_404(queryset, pk=pk)
        if not sertifikasi.is_validated:
            sertifikasi.is_validated = True
            sertifikasi.save()
        return Response(status=status.HTTP_200_OK)

# Filtering API
class SertifikasiList(ListAPIView):
    permission_classes = (IsAuthenticated, )
    serializer_class = SertifikasiSerializer
    queryset = Sertifikasi.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['nama_sertifikasi', 'lembaga_sertifikasi', 'nama', 'tanggal', 'nomor','departemen']


# Export Excel API
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_data(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}-sertifikasi.xlsx'.format(date=datetime.now().strftime('%d-%m-%Y'))

    if request.user.is_admin:
        queryset = Sertifikasi.objects.all()
    else:
        queryset = Sertifikasi.objects.filter(is_validated=True)

    param = request.GET
    if 'tahun' in param:
        queryset = queryset.filter(tanggal__year=param['tahun'])
    if 'departemen' in param:
        queryset = queryset.filter(departemen=param['departemen'])

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Sertifikasi'

    # Define the titles for columns
    columns = [
        'Nama_sertifikasi', 
        'lembaga_sertifikasi', 
        'Nama', 
        'Tanggal', 
        'Nomor',
        'Departemen'
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all kuliah tamu data
    for data in queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            data.nama_sertifikasi,
            data.lembaga_sertifikasi,
            data.nama,
            data.tanggal,
            data.nomor,
            data.departemen
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response