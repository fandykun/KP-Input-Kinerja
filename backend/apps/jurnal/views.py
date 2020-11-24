from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import get_object_or_404

from .models import Jurnal
from .serializers import JurnalSerializer
from ..authentication.permissions import isAdminPermission

# Create your views here.
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListAPIView
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes

from url_filter.integrations.drf import DjangoFilterBackend

from datetime import datetime
from openpyxl import Workbook

class JurnalViewSet(viewsets.ViewSet):
    parser_classes = (MultiPartParser, FormParser)

    def get_all_objects(self, request):
        if request.user.is_admin:
            queryset = Jurnal.objects.all()
        else:
            queryset = Jurnal.objects.filter(is_validated=True)
        return queryset

    def get_permissions(self):
        if self.action == 'destroy':
            permission_classes = [isAdminPermission]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def list(self, request):
        jurnals = self.get_all_objects(request)
        serializer = JurnalSerializer(jurnals, many=True)
        return Response(serializer.data)

    def create(self, request):
        serializer = JurnalSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, pk=None):
        queryset = self.get_all_objects(request)
        jurnal = get_object_or_404(queryset, pk=pk)
        serializer = JurnalSerializer(jurnal)
        return Response(serializer.data)

    def update(self, request, pk=None):
        pass

    def partial_update(self, request, pk=None):
        pass

    def destroy(self, request, pk=None):
        queryset = Jurnal.objects.all()
        jurnal = get_object_or_404(queryset, pk=pk)
        jurnal.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

# Validated API
@api_view(['POST'])
@permission_classes([isAdminPermission])
def set_validate(request, pk):
    if request.method == 'POST':
        queryset = Jurnal.objects.all()
        jurnal = get_object_or_404(queryset, pk=pk)
        if not jurnal.is_validated:
            jurnal.is_validated = True
            jurnal.save()
        return Response(status=status.HTTP_200_OK)
    
# Filtering API
class JurnalList(ListAPIView):
    permission_classes = (IsAuthenticated, )
    serializer_class = JurnalSerializer
    queryset = Jurnal.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = [
        'judul', 
        'author', 
        'published_at',
        'jurnal_vol',
        'jurnal_no',
        'url', 
        'tahun',
        'tingkat',
        'pi',
        'pn',
        'scopus',
        'departemen'
    ]

# Export Excel API
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_data(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}-jurnal.xlsx'.format(date=datetime.now().strftime('%d-%m-%Y'))

    if request.user.is_admin:
        queryset = Jurnal.objects.all()
    else:
        queryset = Jurnal.objects.filter(is_validated=True)

    param = request.GET
    if 'tahun' in param:
        queryset = queryset.filter(tahun=param['tahun'])
    if 'departemen' in param:
        queryset = queryset.filter(departemen=param['departemen'])

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Jurnal'

    # Define the titles for columns
    columns = [
        'Judul', 
        'Author', 
        'Dipublish di',
        'Volume',
        'Nomor',
        'Url', 
        'Tahun',
        'Tingkat',
        'pi',
        'pn',
        'Scopus',
        'Departemen'
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all kuliah tamu data
    for data in queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            data.judul,
            data.author,
            data.published_at,
            data.jurnal_vol,
            data.jurnal_no,
            data.url,
            data.tahun,
            data.tingkat,
            int(data.pi),
            int(data.pn),
            int(data.scopus),
            data.departemen.nama
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response