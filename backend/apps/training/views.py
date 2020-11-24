from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import get_object_or_404

from .models import Training
from .serializers import TrainingSerializer
from ..authentication.permissions import isAdminPermission

from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListAPIView
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes

from url_filter.integrations.drf import DjangoFilterBackend

from datetime import datetime
from openpyxl import Workbook

# Create your views here.
class TrainingAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        trains_dosen = Training.objects.all()
        serializer = TrainingSerializer(trains_dosen, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = TrainingSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TrainingDetailsAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = (IsAuthenticated, )

    def get_object(self, pk):
        try:
            return Training.objects.get(pk=pk)
        except Training.DoesNotExist:
            raise NotFound

    def get(self, request, pk):
        train_dosen = self.get_object(pk)
        serializer = TrainingSerializer(train_dosen)
        return Response(serializer.data)

    def put(self, request, pk):
        train_dosen = self.get_object(pk)
        serializer = TrainingSerializer(train_dosen, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        train_dosen = self.get_object(pk)
        if isAdminPermission:
            train_dosen.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

# Validated API
@api_view(['POST'])
@permission_classes([isAdminPermission])
def set_validate(request, pk):
    if request.method == 'POST':
        queryset = Training.objects.all()
        train = get_object_or_404(queryset, pk=pk)
        if not train.is_validated:
            train.is_validated = True
            train.save()
        return Response(status=status.HTTP_200_OK)

class TrainingList(ListAPIView):
    permission_classes = (IsAuthenticated, )
    serializer_class = TrainingSerializer
    queryset = Training.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = [
        'judul',
        'peserta',
        'jenis',
        'date_start',
        'date_end',
        'tempat',
        'departemen',
    ]


# Export Excel API
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def export_data(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}-training.xlsx'.format(date=datetime.now().strftime('%d-%m-%Y'))

    # if request.user.is_admin:
    queryset = Training.objects.all()
    # else:
    #     queryset = Training.objects.filter(is_validated=True)

    param = request.GET
    if 'tahun' in param:
        queryset = queryset.filter(date_end__year=param['tahun'])
    if 'departemen' in param:
        queryset = queryset.filter(departemen=param['departemen'])

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Training'

    # Define the titles for columns
    columns = [
        'Judul',
        'Peserta',
        'Jenis',
        'Tempat',
        'Tanggal Mulai',
        'Tanggal Berakhir',
        'Departemen',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all kuliah tamu data
    for data in queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            data.judul,
            data.peserta,
            data.jenis,
            data.tempat,
            data.date_start,
            data.date_end,
            data.departemen.nama
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response